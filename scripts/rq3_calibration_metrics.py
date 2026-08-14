"""Shared RQ3 loader and pacing-calibration generator.

The current summary pipeline imports this module's source definitions, workbook
loader, and inclusive-percentile helpers. It also emits the retained
data/rq3/calibration/ diagnostics. The core definitions are:

    d  = ceil( max(0, Bhat + 2*Chat - max(0,T)) / 2 )   the deployed pacer
    d* = ceil( max(0, B    + 2*C    - max(0,T)) / 2 )   the same condition on
                                                        measured inputs

with Bhat = PacingReplay.beforeBacklogMs, Chat = beforeDraftSequenceReservedDurationMs,
B = RQ3Pacing.realBacklogMs, C = PacingReplay.draftSequenceDurationMs,
T = PacingReplay.beforeTimeToDeadlineMs.

Run:
    python scripts/rq3_calibration_metrics.py sampling
    python scripts/rq3_calibration_metrics.py raw --no-write   # old source, for deltas
"""
import csv
import math
import os
import sys
import warnings

import openpyxl

warnings.filterwarnings('ignore')

PAPER = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ML = os.path.join(os.path.dirname(PAPER), 'ML')
OUT = os.path.join(PAPER, 'data', 'rq3', 'calibration')

CONDITIONS = ('12mp_normal', '24mp_memory')
LABEL = {'12mp_normal': '12MP normal', '24mp_memory': '24MP mode + memory pressure'}

SOURCES = {
    # The set the committed revision was built from: one workbook per condition,
    # 43 and 39 eligible runs.  Kept so the change against it can be quantified.
    'raw': {
        '12mp_normal': [f'{ML}/data/0803_FULL/SM-S948U_metrics_12MP_normal_0803.xlsx'],
        '24mp_memory': [f'{ML}/data/0803_FULL/SM-S948U_metrics_24MP_memory_0803.xlsx'],
    },
    'original': {
        c: [f'{PAPER}/data/ablation_original/48U_metrics_{n}_0803_{i}.xlsx' for i in (1, 2)]
        for c, n in (('12mp_normal', '12MP_normal'), ('24mp_memory', '24MP_memory'))
    },
    # Current source of record: the RQ1(a) full arm balanced to ten runs per
    # (condition, starting level) cell.  See data/ablation_sampling/README.md.
    'sampling': {
        c: [f'{PAPER}/data/ablation_sampling/48U_metrics_{n}_0803_{i}.xlsx' for i in (1, 2)]
        for c, n in (('12mp_normal', '12MP_normal'), ('24mp_memory', '24MP_memory'))
    },
}

BANDS = ('Lv0--2', 'Lv3--4', 'Lv5--6')


def band_of(level):
    return BANDS[0] if level <= 2 else BANDS[1] if level <= 4 else BANDS[2]


def read(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {}
    for i, h in enumerate(header):
        if h and h not in idx:
            idx[h] = i
    rows = [{k: (r[i] if len(r) > i else None) for k, i in idx.items()}
            for r in it if any(c is not None for c in r)]
    wb.close()
    return rows


def truthy(v):
    return str(v).strip().lower() in ('true', '1', '1.0')


def pct_inc(values, q):
    """Excel PERCENTILE.INC, as in docs/rq-evidence.md (Part 2) section 3.4."""
    v = sorted(float(x) for x in values if x is not None)
    if not v:
        return None
    rank = (len(v) - 1) * q
    lo, hi = math.floor(rank), math.ceil(rank)
    return v[lo] if lo == hi else v[lo] + (v[hi] - v[lo]) * (rank - lo)


def med(values):
    return pct_inc(values, 0.5)


def load(condition, files):
    """Analyzed transitions and the run-eligibility audit for one condition.

    Run eligibility: complete valid 30-shot runs. Timeout-labelled records in
    this collection are known measurement errors and are removed as invalid
    observations, not actual timeout outcomes. In the 24MP condition a run
    captures MP24 on shots 1-2 and MP12 from shot 3; from Lv5 it falls back to
    MP12 for the whole run, which is production behaviour and belongs to the
    condition, so an MP12-from-shot-1 run is excluded only below that level.
    Transition eligibility: shots 2-30 with a recorded pacing decision and a
    complete prior-Draft timeline. Watchdog-truncated captures are dropped as
    transitions because a truncated node has no clean executed-Draft duration,
    while their runs are retained.
    """
    transitions, audit = [], []
    for part, path in enumerate(files, start=1):
        pacing = read(path, 'RQ3Pacing')
        replay = {r['captureIndex']: r for r in read(path, 'PacingReplay')}
        # C_mand: the suffix admission can never skip -- DynamicFunction plus the
        # encoding pass -- plus the non-node Draft overhead.
        dyn = {r['captureIndex']: r['durationMs'] for r in read(path, 'DynamicFunctionNode')}
        enc = {r['captureIndex']: r['durationMs'] for r in read(path, 'SecImageCodecNode')
               if str(r['workloadKey']).startswith('ENCODING')}
        by_run = {}
        for r in pacing:
            by_run.setdefault(r['runId'], []).append(r)
        for summary in read(path, 'RQ3Summary'):
            run = f'{part}#{int(summary["runId"])}'
            shots = sorted(by_run.get(summary['runId'], []), key=lambda r: r['runShotIndex'])
            level = int(summary['startingOverheatLevel'])
            excluded = None
            if not truthy(summary['isComplete30ShotRun']):
                excluded = 'incomplete'
            elif any(truthy(r['captureTimedOut']) for r in shots):
                excluded = 'invalid-timeout-measurement'
            elif condition == '24mp_memory' and level <= 4:
                head = [r['sizeBucket'] for r in shots if r['runShotIndex'] in (1, 2)]
                if head and all(b == 'MP12' for b in head):
                    excluded = 'mp12-from-shot-1'
            audit.append({'run': run, 'condition': LABEL[condition], 'level': level,
                          'shots': len(shots), 'excluded': excluded or '',
                          'included': excluded is None})
            if excluded:
                continue
            for r in shots:
                if not 2 <= r['runShotIndex'] <= 30:
                    continue
                if not (truthy(r['pacingDecisionRecorded'])
                        and truthy(r['realTraceCompleteBeforeDelay'])):
                    continue
                pr = replay.get(r['captureIndex'])
                if pr is None:
                    continue
                d, B = pr['beforeAppliedDelayMs'], r['realBacklogMs']
                C, T = pr['draftSequenceDurationMs'], pr['beforeTimeToDeadlineMs']
                Bhat, Chat = pr['beforeBacklogMs'], pr['beforeDraftSequenceReservedDurationMs']
                if any(v is None for v in (d, B, C, T, Bhat, Chat)):
                    continue
                d = float(d)
                required = float(math.ceil(max(0.0, B + 2 * C - max(0.0, T)) / 2))
                mandatory = None
                if (dyn.get(r['captureIndex']) is not None
                        and enc.get(r['captureIndex']) is not None
                        and pr['workloadSequenceDurationMs'] is not None):
                    c_mand = (dyn[r['captureIndex']] + enc[r['captureIndex']]
                              + (C - pr['workloadSequenceDurationMs']))
                    mandatory = float(math.ceil(max(0.0, B + 2 * c_mand - max(0.0, T)) / 2))
                transitions.append({
                    'run': run, 'shot': int(r['runShotIndex']), 'level': level,
                    'band': band_of(level), 'size': r['sizeBucket'],
                    'd': d, 'required': required, 'error': d - required,
                    'mandatory': mandatory, 'shotToShot': r['shotToShotTimeMs'],
                    'B': float(B), 'C': float(C), 'T': float(T),
                    'Bhat': float(Bhat), 'Chat': float(Chat),
                    'reserve': float(Chat) - float(C),
                    'backlog': (float(Bhat) - float(B)) / 2,
                    'budget': float(pr['captureTimeoutMs']),
                    'margin': r['timeoutMarginMs'],
                    'queueWait': pr['realQueueWaitMs'],
                    'queueDepth': r['realQueueDepth'],
                    'crossSize': pr['draftSequenceReserveCrossSizeContaminationMs'],
                    'watchdog': truthy(r['captureWatchdogFailed']),
                    'executed': r['executedWorkloadSequenceKey'],
                    'demoted': r['plannedWorkloadSequenceKey'] != r['executedWorkloadSequenceKey'],
                })
    kept = [t for t in transitions if not t['watchdog']]
    return kept, transitions, audit


def analyse(condition, files):
    tx, tx_all, audit = load(condition, files)
    budget = med([t['budget'] for t in tx])
    required = [t for t in tx if t['required'] > 0]
    paced = [t for t in tx if t['d'] > 0]
    inplay = [t for t in tx if t['d'] > 0 or t['required'] > 0]

    rows = []
    for band in BANDS + ('All',):
        sel = (lambda p: p) if band == 'All' else (lambda p: [t for t in p if t['band'] == band])
        req_b, paced_b, tx_b = sel(required), sel(paced), sel(tx)
        if not req_b:
            continue
        covered = [t for t in req_b if t['d'] >= t['required']]
        deficits = [t['required'] - t['d'] for t in req_b if t['d'] < t['required']]
        req_paced = [t for t in req_b if t['d'] > 0]
        rows.append({
            'band': band, 'nreq': len(req_b), 'ncovered': len(covered),
            'coverage': 100 * len(covered) / len(req_b),
            'worst': max(deficits) if deficits else None,
            'surplus': med([t['error'] for t in paced_b]),
            'reserveTerm': med([t['reserve'] for t in req_paced]),
            'backlogTerm': med([t['backlog'] for t in req_paced]),
            'backlogP95': 100 * pct_inc([t['B'] for t in tx_b], 0.95) / budget,
        })

    covered_all = [t for t in required if t['d'] >= t['required']]
    zero_paced = [t for t in paced if t['required'] == 0]
    burst = {}
    for t in tx:
        burst.setdefault(t['run'], []).append(t)

    return {
        'condition': condition, 'tx': tx, 'audit': audit, 'budget': budget,
        'required': required, 'paced': paced, 'inplay': inplay, 'rows': rows,
        'runs': sum(1 for a in audit if a['included']),
        'runsExcluded': [a for a in audit if not a['included']],
        'nAnalyzed': len(tx), 'nAnalyzedWithWatchdog': len(tx_all),
        'nWatchdog': sum(1 for t in tx_all if t['watchdog']),
        'pacedPercent': 100 * len(paced) / len(tx),
        'coverage': 100 * len(covered_all) / len(required),
        'nZeroPaced': len(zero_paced),
        'zeroPacedPercent': 100 * len(zero_paced) / len(paced),
        'surplusP95': pct_inc([t['error'] for t in paced], 0.95),
        'requiredP50': med([t['required'] for t in required]),
        'requiredP95': pct_inc([t['required'] for t in required], 0.95),
        'requiredMax': max(t['required'] for t in required),
        'appliedMax': max(t['d'] for t in required),
        'burstSurplusP50': med([sum(t['error'] for t in v) for v in burst.values()]),
        'burstDelayP50': med([sum(t['d'] for t in v) for v in burst.values()]),
        'underPercentAll': 100 * sum(1 for t in tx if t['error'] < 0) / len(tx),
        'zeroErrorPercentAll': 100 * sum(1 for t in tx if t['error'] == 0) / len(tx),
        'underPercentInplay': 100 * sum(1 for t in inplay if t['error'] < 0) / len(inplay),
        'slackP5': pct_inc([t['margin'] for t in tx if t['margin'] is not None], 0.05),
        'slackMin': min(t['margin'] for t in tx if t['margin'] is not None),
        'demoted': sum(1 for t in required if t['demoted']),
        'errorShare': sorted(100 * t['error'] / t['budget'] for t in inplay),
        'zeroPaced': zero_paced,
    }


def rank(values):
    order = sorted(range(len(values)), key=lambda i: values[i])
    r = [0.0] * len(values)
    i = 0
    while i < len(order):
        j = i
        while j + 1 < len(order) and values[order[j + 1]] == values[order[i]]:
            j += 1
        avg = (i + j) / 2 + 1
        for k in range(i, j + 1):
            r[order[k]] = avg
        i = j + 1
    return r


def corr(xs, ys):
    n = len(xs)
    mx, my = sum(xs) / n, sum(ys) / n
    sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    sxx = sum((x - mx) ** 2 for x in xs)
    syy = sum((y - my) ** 2 for y in ys)
    r = sxy / math.sqrt(sxx * syy) if sxx and syy else float('nan')
    slope = sxy / sxx if sxx else float('nan')
    return r, slope, my - slope * mx


def notes(res):
    """The quantities the .tex header comments quote but the table does not print."""
    tx, required, paced = res['tx'], res['required'], res['paced']
    x = [t['required'] for t in required]
    y = [t['d'] for t in required]
    r, slope, intercept = corr(x, y)
    rs, _, _ = corr(rank(x), rank(y))
    print(f'  envelope, not tracking: Pearson r {r:+.3f}  Spearman {rs:+.3f}  '
          f'LS fit d = {slope:.2f} d* + {intercept:.0f} ms')
    print(f'    d* quartiles {pct_inc(x,0.25):.0f} / {pct_inc(x,0.5):.0f} / {pct_inc(x,0.75):.0f}'
          f'   d quartiles {pct_inc(y,0.25):.0f} / {pct_inc(y,0.5):.0f} / {pct_inc(y,0.75):.0f}')
    low = [t['d'] for t in required if t['required'] <= 100]
    high = [t['d'] for t in required if t['required'] > 300]
    print(f'    coarse response: d* <= 100 ms -> median applied {med(low):.0f} ms (n={len(low)});'
          f'  d* > 300 ms -> {med(high):.0f} ms (n={len(high)})' if high else
          f'    coarse response: d* <= 100 ms -> median applied {med(low):.0f} ms (n={len(low)})')

    clear = sum(1 for t in required if t['error'] > 100)
    under = [t for t in required if t['d'] < t['required']]
    near = sum(1 for t in under if t['required'] - t['d'] <= 100)
    print(f'  scatter guides: {100*clear/len(required):.0f}% clear the diagonal by > 100 ms;'
          f'  {len(under)} below it, {len(under)-near} of them beyond -100 ms')
    print(f'    plotted range: max d* {max(x):.0f} ms, max d {max(y):.0f} ms, '
          f'{sum(1 for v in y if v > 1500)} point(s) above y = 1500')
    if under:
        print(f'    smallest realized margin at an under-paced transition '
              f'{min(t["margin"] for t in under if t["margin"] is not None):.0f} ms')

    # The mandatory floor, reported in the caption because it does not vary.
    mand = [t for t in tx if t['mandatory'] is not None]
    if mand:
        pos = [t for t in mand if t['mandatory'] > 0]
        breached = [t for t in mand if t['d'] < t['mandatory']]
        print(f'  mandatory floor d*_mand: positive on {len(pos)} of {len(mand)} transitions, '
              f'breached on {len(breached)}'
              + (f'; covered pairs (applied vs floor) '
                 + ', '.join(f'{t["d"]:.0f} vs {t["mandatory"]:.0f}' for t in pos) if pos else ''))
    fully = [t for t in tx if str(t['executed']).count('>') == 1]
    print(f'  executed sequence demoted to DYNAMIC_FUNCTION>ENCODING on {len(fully)} transitions')
    s2s = [t['shotToShot'] for t in tx if t['shotToShot']]
    print(f'  median shot-to-shot interval {med(s2s):.0f} ms')

    # Surplus decomposition, exact away from the max(0,.) clip.
    both = [t for t in required if t['d'] > 0]
    resid = [t['error'] - (t['reserve'] + t['backlog']) for t in both]
    print(f'  decomposition check on {len(both)} transitions with d > 0 and d* > 0: '
          f'residual mean {sum(resid)/len(resid):+.2f} ms, max |residual| {max(abs(v) for v in resid):.2f} ms')

    # Demotion split: the first objection to the surplus.
    dem = [t['error'] for t in required if t['demoted']]
    und = [t['error'] for t in required if not t['demoted']]
    dcov = [t for t in required if t['demoted'] and t['d'] >= t['required']]
    ucov = [t for t in required if not t['demoted'] and t['d'] >= t['required']]
    try:
        from scipy.stats import mannwhitneyu
        p = f'{mannwhitneyu(dem, und).pvalue:.2f}'
    except ImportError:
        p = 'n/a'
    below_d = 100 * sum(1 for t in required if t['demoted'] and t['d'] < t['required']) / len(dem)
    below_u = 100 * sum(1 for t in required if not t['demoted'] and t['d'] < t['required']) / len(und)
    print(f'  demotion split: {len(dem)} demoted / {len(und)} undemoted; '
          f'coverage {100*len(dcov)/len(dem):.1f}% vs {100*len(ucov)/len(und):.1f}%; '
          f'median surplus {med(dem):+.0f} vs {med(und):+.0f} ms; '
          f'below diagonal {below_d:.1f}% vs {below_u:.1f}%; Mann-Whitney p = {p}')

    # Cross-size contamination, the alternative explanation for the 24MP reserve.
    contam = [t for t in paced if t['crossSize'] and t['crossSize'] > 0]
    if contam:
        clean = [t for t in paced if not (t['crossSize'] and t['crossSize'] > 0)]
        print(f'  cross-size contamination on {100*len(contam)/len(paced):.1f}% of paced; '
              f'median reserve term {med([t["reserve"] for t in contam]):.0f} vs '
              f'{med([t["reserve"] for t in clean]):.0f} ms (contaminated vs clean)')

    # The d* = 0 column of the figure: not an idle pipeline.
    z = res['zeroPaced']
    waits = [t['queueWait'] for t in z if t['queueWait'] is not None]
    depth = [t['queueDepth'] for t in z if t['queueDepth'] is not None]
    ratio = [100 * t['d'] / t['queueWait'] for t in z if t['queueWait']]
    print(f'  zero-requirement paced n={len(z)}: realQueueWait P10/P50/P90 '
          f'{pct_inc(waits,0.1):.0f}/{pct_inc(waits,0.5):.0f}/{pct_inc(waits,0.9):.0f} ms; '
          f'queueDepth >= 1 on {100*sum(1 for v in depth if v >= 1)/len(depth):.1f}% (median {med(depth):.0f}); '
          f'delay covers median {med(ratio):.0f}% of the wait, below it on '
          f'{100*sum(1 for v in ratio if v < 100)/len(ratio):.1f}%')


def emit(res):
    """Figure inputs, in the formats the .tex files read."""
    os.makedirs(OUT, exist_ok=True)
    cond = res['condition']

    def write(name, header, rows):
        with open(os.path.join(OUT, name), 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)

    for suffix, keep in (('req_kept', False), ('req_changed', True)):
        write(f'scatter_{cond}_{suffix}.csv', ['required_ms', 'applied_ms'],
              [[t['required'], t['d']] for t in res['required'] if t['demoted'] is keep])
    write(f'scatter_{cond}_zero.csv', ['required_ms', 'applied_ms'],
          [[0.0, t['d']] for t in res['paced'] if t['required'] == 0])

    for suffix, keep in (('unchanged', False), ('admission_changed', True)):
        write(f'{cond}_{suffix}.csv',
              ['run', 'level', 'shot', 'actual_size', 'required_delay_ms',
               'applied_delay_ms', 'calibration_error_ms'],
              [[t['run'], t['level'], t['shot'], t['size'], t['required'], t['d'], t['error']]
               for t in res['required'] if t['demoted'] is keep])

    # ECDF over the in-play population, x as a share of the Capture Timeout budget.
    values, n, rows = res['errorShare'], len(res['errorShare']), []
    for i, v in enumerate(values, start=1):
        if i == n or values[i] != v:
            rows.append([round(v, 4), round(i / n, 5)])
    write(f'error_ecdf_{cond}.csv', ['error_pct', 'cdf'], rows)


def write_audit(results):
    with open(os.path.join(OUT, 'run_audit.csv'), 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['runKey', 'condition', 'startingLevel', 'shots', 'excludedReason', 'included'])
        for res in results:
            for a in res['audit']:
                w.writerow([f'{LABEL[res["condition"]]}#{a["run"]}', a['condition'],
                            a['level'], a['shots'], a['excluded'], a['included']])


def write_summary(results):
    with open(os.path.join(OUT, 'summary.csv'), 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['condition', 'band', 'nRequired', 'nCovered', 'coveragePercent',
                    'worstDeficitMs', 'surplusP50Ms', 'reserveTermMs', 'backlogTermMs',
                    'backlogP95Percent'])
        for res in results:
            for r in res['rows']:
                w.writerow([LABEL[res['condition']], r['band'], r['nreq'], r['ncovered'],
                            round(r['coverage'], 1),
                            '' if r['worst'] is None else round(r['worst']),
                            round(r['surplus']), round(r['reserveTerm']),
                            round(r['backlogTerm']), round(r['backlogP95'], 1)])


def report(res):
    print(f'=== {LABEL[res["condition"]]}')
    reasons = {}
    for a in res['runsExcluded']:
        reasons[a['excluded']] = reasons.get(a['excluded'], 0) + 1
    print(f'  runs {res["runs"]} eligible of {res["runs"] + len(res["runsExcluded"])} '
          f'(excluded: {reasons or "none"})')
    print(f'  analyzed transitions {res["nAnalyzed"]} '
          f'({res["nAnalyzedWithWatchdog"]} before dropping {res["nWatchdog"]} watchdog-truncated)')
    print(f'  paced {len(res["paced"])} ({res["pacedPercent"]:.1f}%)  '
          f'required>0 {len(res["required"])}  coverage {res["coverage"]:.1f}%')
    print(f'  {"band":6s} {"nreq":>5s} {"cov%":>7s} {"covered":>9s} {"worst":>6s} '
          f'{"surp":>6s} {"resv":>6s} {"back":>6s} {"bkP95":>6s}')
    for r in res['rows']:
        print(f'  {r["band"]:6s} {r["nreq"]:5d} {r["coverage"]:6.1f}% '
              f'{r["ncovered"]:4d} of {r["nreq"]:<3d} '
              f'{"--" if r["worst"] is None else round(r["worst"]):>6} '
              f'{round(r["surplus"]):>6} {round(r["reserveTerm"]):>+6} '
              f'{round(r["backlogTerm"]):>+6} {r["backlogP95"]:5.1f}%')
    e = res['errorShare']
    print(f'  surplus P95 {res["surplusP95"]:.0f} ms   '
          f'burst surplus P50 {res["burstSurplusP50"]/1000:.2f} s  '
          f'(burst total delay {res["burstDelayP50"]/1000:.2f} s)')
    print(f'  required delay d*>0: P50 {res["requiredP50"]:.0f}  P95 {res["requiredP95"]:.0f}  '
          f'max {res["requiredMax"]:.0f}   applied max {res["appliedMax"]:.0f}')
    print(f'  calibration error over all transitions: P(<0) {res["underPercentAll"]:.1f}%  '
          f'P(=0) {res["zeroErrorPercentAll"]:.1f}%')
    print(f'  in-play n={len(res["inplay"])} (% of budget): min {e[0]:+.1f}  '
          f'P50 {pct_inc(e,0.5):+.1f}  P90 {pct_inc(e,0.9):+.1f}  P95 {pct_inc(e,0.95):+.1f}  '
          f'max {e[-1]:+.1f}   under-paced {res["underPercentInplay"]:.1f}%')
    print(f'  zero-requirement paced {res["nZeroPaced"]} ({res["zeroPacedPercent"]:.1f}% of paced)')
    print(f'  realized slack P5 {res["slackP5"]:.0f} ms  min {res["slackMin"]:.0f} ms  '
          f'budget {res["budget"]:.0f} ms')
    print(f'  demoted after decision: {res["demoted"]} of {len(res["required"])} required')
    notes(res)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('-')]
    source = args[0] if args else 'sampling'
    write = '--no-write' not in sys.argv[1:]
    results = [analyse(c, SOURCES[source][c]) for c in CONDITIONS]
    print(f'source: {source}\n')
    for res in results:
        report(res)
        print()
    if write:
        for res in results:
            emit(res)
        write_audit(results)
        write_summary(results)
        print(f'wrote {OUT}')


if __name__ == '__main__':
    main()
