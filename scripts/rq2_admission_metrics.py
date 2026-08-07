"""RQ2 admission-decision quality: recompute tables/tab_rq2_admission_summary.tex
from an arbitrary workbook set.

Follows docs/rq1-rq3-metrics-guide.md sections 3.2 (run reconstruction),
3.5 (valid-run policy / dedup) and 5.2-5.3 (controller-enforced metrics and the
always-admit audit).  Every source set below goes through identical logic, so a
difference between two printed blocks is a data difference, not a method
difference.

Run:  python scripts/rq2_admission_metrics.py
"""
import openpyxl, warnings, os, statistics, sys

warnings.filterwarnings('ignore')

PAPER = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ML = os.path.join(os.path.dirname(PAPER), 'ML')

ORIG = os.path.join(PAPER, 'data', 'ablation_original')
SAMP = os.path.join(PAPER, 'data', 'ablation_sampling')
RQ2_0729 = os.path.join(ML, 'data', '0729_RQ2')
FULL_0803 = os.path.join(ML, 'data', '0803_FULL')

# Source sets.  Each entry: label -> {condition -> [workbook paths]}
ENFORCED = {
    'A0  0729 Full campaign': {
        '12MP normal': [os.path.join(RQ2_0729, '48U_metrics_12MP_normal_0729.xlsx')],
        '24MP memory': [os.path.join(RQ2_0729, '48U_metrics_24MP_memory_0729.xlsx')],
    },
    'A1  current  ML/0803_FULL single export': {
        '12MP normal': [os.path.join(FULL_0803, 'SM-S948U_metrics_12MP_normal_0803.xlsx')],
        '24MP memory': [os.path.join(FULL_0803, 'SM-S948U_metrics_24MP_memory_0803.xlsx')],
    },
    'A2  data/ablation_original  Full (_1 + _2)': {
        '12MP normal': [os.path.join(ORIG, '48U_metrics_12MP_normal_0803_1.xlsx'),
                        os.path.join(ORIG, '48U_metrics_12MP_normal_0803_2.xlsx')],
        '24MP memory': [os.path.join(ORIG, '48U_metrics_24MP_memory_0803_1.xlsx'),
                        os.path.join(ORIG, '48U_metrics_24MP_memory_0803_2.xlsx')],
    },
    'A3  data/ablation_sampling  Full (_1 + _2)': {
        '12MP normal': [os.path.join(SAMP, '48U_metrics_12MP_normal_0803_1.xlsx'),
                        os.path.join(SAMP, '48U_metrics_12MP_normal_0803_2.xlsx')],
        '24MP memory': [os.path.join(SAMP, '48U_metrics_24MP_memory_0803_1.xlsx'),
                        os.path.join(SAMP, '48U_metrics_24MP_memory_0803_2.xlsx')],
    },
}

AUDIT = {
    'B1  current  0729 PacingOnly (_1 + _2)': {
        '12MP normal': [os.path.join(RQ2_0729, '48U_metrics_12MP_normal_0729_PacingOnly_1.xlsx'),
                        os.path.join(RQ2_0729, '48U_metrics_12MP_normal_0729_PacingOnly_2.xlsx')],
        '24MP memory': [os.path.join(RQ2_0729, '48U_metrics_24MP_memory_0729_PacingOnly_1.xlsx'),
                        os.path.join(RQ2_0729, '48U_metrics_24MP_memory_0729_PacingOnly_2.xlsx')],
    },
    'B2  data/ablation_sampling  pacing_only_0803': {
        '12MP normal': [os.path.join(SAMP, '48U_metrics_12MP_normal_pacing_only_0803.xlsx')],
        '24MP memory': [os.path.join(SAMP, '48U_metrics_24MP_memory_pacing_only_0803.xlsx')],
    },
    'B3  data/ablation_sampling  baseline_0803 (no control)': {
        '12MP normal': [os.path.join(SAMP, '48U_metrics_12MP_normal_baseline_0803.xlsx')],
        '24MP memory': [os.path.join(SAMP, '48U_metrics_24MP_memory_baseline_0803.xlsx')],
    },
}

# Predeclared manifest exclusion for the 0729 audit set: source run 16 of the
# 24MP workbook 1 was invalid/incomplete (guide section 5.3).
MANIFEST_DROP = {
    ('48U_metrics_24MP_memory_0729_PacingOnly_1.xlsx', 16),
}

GROUP = {'Bokeh': 'Multi-frame', 'Filter': 'Single-frame'}


def read_sheet(path, name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close()
        return None, []
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    idx = {}
    for i, h in enumerate(hdr):
        if h and h not in idx:
            idx[h] = i
    rows = [r for r in it if any(c is not None for c in r)]
    wb.close()
    return idx, rows


def cell(r, i):
    return r[i] if (i is not None and len(r) > i) else None


def truthy(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    return str(v).strip().lower() in ('true', 'yes', '1', 'y')


def num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def load_decisions(paths, max_shot=30):
    """Return the capture-level selected admission decisions of a workbook set.

    One Bokeh row and one Filter row per capture, runs delimited by a
    ppSequenceId reset, shots after `max_shot` dropped, identical run
    signatures counted once across workbooks.
    """
    out, seen = [], set()
    stats = dict(runs_seen=0, runs_dup=0, runs_kept=0, runs_manifest=0,
                 captures=0, short_runs=0)
    for p in paths:
        if not os.path.exists(p):
            print(f'  !! missing {p}')
            continue
        base = os.path.basename(p)
        ai, arows = read_sheet(p, 'AdmissionReplay')
        pi, prows = read_sheet(p, 'PacingReplay')
        draft_end = {}
        for r in prows:
            c = cell(r, pi['captureIndex'])
            if c is not None:
                draft_end[c] = num(cell(r, pi['draftEndUptimeMs']))

        # captures in worksheet order, run split on ppSequenceId reset
        order, pp = [], {}
        for r in arows:
            c = cell(r, ai['captureIndex'])
            if c not in pp:
                order.append(c)
                pp[c] = cell(r, ai['ppSequenceId'])
        runs, cur, prev = [], [], None
        for c in order:
            if cur and prev is not None and pp[c] is not None and pp[c] <= prev:
                runs.append(cur)
                cur = []
            cur.append(c)
            prev = pp[c]
        if cur:
            runs.append(cur)

        by_cap = {}
        for r in arows:
            by_cap.setdefault(cell(r, ai['captureIndex']), []).append(r)

        for k, run in enumerate(runs, start=1):
            stats['runs_seen'] += 1
            if (base, k) in MANIFEST_DROP:
                stats['runs_manifest'] += 1
                continue
            sig = tuple((pp[c], cell(by_cap[c][0], ai['nodeStartUptimeMs']),
                         cell(by_cap[c][0], ai['timeoutDeadlineUptimeMs'])) for c in run)
            if sig in seen:
                stats['runs_dup'] += 1
                continue
            seen.add(sig)
            stats['runs_kept'] += 1
            if len(run) < max_shot:
                stats['short_runs'] += 1
            for shot, c in enumerate(run[:max_shot], start=1):
                stats['captures'] += 1
                rows = sorted(by_cap[c], key=lambda r: (cell(r, ai['nodeOrder']) or 0))
                for stage, group in GROUP.items():
                    d = next((r for r in rows if cell(r, ai['admissionStage']) == stage), None)
                    if d is None:
                        continue
                    b = num(cell(d, ai['beforeBudgetMs']))
                    start = num(cell(d, ai['nodeStartUptimeMs']))
                    end = draft_end.get(c)
                    cost = end - start if (end is not None and start is not None) else None
                    out.append(dict(
                        book=base, run=k, shot=shot, capture=c, group=group,
                        budget=b, cost=cost,
                        watchdog=truthy(cell(d, ai['beforeWatchdogTimedOut'])) or
                                 truthy(cell(d, ai['beforeCaptureWatchdogFailed'])),
                        timed_out=truthy(cell(d, ai['beforeCaptureTimedOut'])),
                        eff_admit=truthy(cell(d, ai['beforeEffectiveAdmit'])),
                        model_admit=truthy(cell(d, ai['afterModelAdmit'])),
                        after_eff_admit=truthy(cell(d, ai['afterEffectiveAdmit'])),
                    ))
    return out, stats


def deadline_of(paths):
    vals = []
    for p in paths:
        if not os.path.exists(p):
            continue
        pi, prows = read_sheet(p, 'PacingReplay')
        if pi and 'captureTimeoutMs' in pi:
            vs = [num(cell(r, pi['captureTimeoutMs'])) for r in prows]
            vs = [v for v in vs if v]
            if vs:
                vals.append(statistics.median(vs))
    return statistics.median(vals) if vals else None


def pct(a, b):
    return 100.0 * a / b if b else float('nan')


def enforced_block(label, sets):
    print(f'\n{"=" * 100}\nCONTROLLER-ENFORCED RUNS   {label}\n{"=" * 100}')
    print(f'{"condition":13s} {"group":13s} {"runs":>5s} {"caps":>5s} '
          f'{"decisions":>9s} {"admits":>7s} {"admit%":>7s} '
          f'{"succ":>6s} {"succ%":>7s} {"unsafe":>6s} {"unsafe%":>7s} '
          f'{"[wd]":>5s} {"[C>B]":>6s} {"noC":>4s}')
    for cond, paths in sets.items():
        dec, st = load_decisions(paths)
        for group in ('Multi-frame', 'Single-frame'):
            g = [d for d in dec if d['group'] == group]
            adm = [d for d in g if d['eff_admit']]
            miss = [d for d in adm if d['cost'] is None]
            wd = [d for d in adm if d['watchdog']]
            over = [d for d in adm if not d['watchdog'] and d['cost'] is not None
                    and d['cost'] > d['budget']]
            unsafe = len(wd) + len(over)
            succ = len(adm) - unsafe - len(miss)
            print(f'{cond:13s} {group:13s} {st["runs_kept"]:5d} {st["captures"]:5d} '
                  f'{len(g):9d} {len(adm):7d} {pct(len(adm), len(g)):7.1f} '
                  f'{succ:6d} {pct(succ, len(adm)):7.1f} '
                  f'{unsafe:6d} {pct(unsafe, len(adm)):7.1f} '
                  f'{len(wd):5d} {len(over):6d} {len(miss):4d}')
        print(f'  ({cond}: {st["runs_seen"]} runs read, {st["runs_dup"]} duplicate, '
              f'{st["runs_manifest"]} manifest-dropped, {st["runs_kept"]} kept, '
              f'{st["short_runs"]} shorter than 30 shots)')


def audit_block(label, sets, decision_field='model_admit'):
    print(f'\n{"=" * 100}\nALWAYS-ADMIT MODEL AUDIT   {label}   (decision = {decision_field})\n{"=" * 100}')
    print(f'{"condition":13s} {"group":13s} {"caps":>5s} '
          f'{"FA":>5s} {"FA%":>6s} {"FS":>4s} {"FS%":>6s} {"margin+":>8s} '
          f'{"UA":>4s} {"UA%":>6s} {"US":>4s} {"US%":>6s} {"overrun-":>8s}')
    for cond, paths in sets.items():
        dec, st = load_decisions(paths)
        D = deadline_of(paths)
        for group in ('Multi-frame', 'Single-frame'):
            g = [d for d in dec if d['group'] == group and d['cost'] is not None]
            feas = [d for d in g if not d['watchdog'] and d['cost'] <= d['budget']]
            unsf = [d for d in g if d['watchdog'] or d['cost'] > d['budget']]
            fa = [d for d in feas if d[decision_field]]
            fs = [d for d in feas if not d[decision_field]]
            ua = [d for d in unsf if d[decision_field]]
            us = [d for d in unsf if not d[decision_field]]
            marg = [100.0 * (d['budget'] - d['cost']) / D for d in fs] if D else []
            over = [100.0 * (d['cost'] - d['budget']) / D for d in us] if D else []
            m50 = statistics.median(marg) if marg else None
            o50 = statistics.median(over) if over else None
            print(f'{cond:13s} {group:13s} {len(g):5d} '
                  f'{len(fa):5d} {pct(len(fa), len(feas)):6.1f} '
                  f'{len(fs):4d} {pct(len(fs), len(feas)):6.1f} '
                  f'{("--" if m50 is None else f"+{m50:.1f}%"):>8s} '
                  f'{len(ua):4d} {pct(len(ua), len(unsf)):6.1f} '
                  f'{len(us):4d} {pct(len(us), len(unsf)):6.1f} '
                  f'{("--" if o50 is None else f"-{o50:.1f}%"):>8s}')
        print(f'  ({cond}: {st["runs_seen"]} runs read, {st["runs_dup"]} duplicate, '
              f'{st["runs_manifest"]} manifest-dropped, {st["runs_kept"]} kept, '
              f'{st["captures"]} captures, {st["short_runs"]} shorter than 30 shots, '
              f'deadline D={D})')


# ---------------------------------------------------------------- comparison

# Published cells of tables/tab_rq2_admission_summary.tex, for the side-by-side.
PUBLISHED_ENFORCED = {
    ('12MP normal', 'Multi-frame'): (828, 1341, 827, 1, 'watchdog'),
    ('12MP normal', 'Single-frame'): (1297, 1340, 1297, 0, ''),
    ('24MP memory', 'Multi-frame'): (554, 1193, 553, 1, 'watchdog'),
    ('24MP memory', 'Single-frame'): (1047, 1192, 1047, 0, ''),
}
# The published Always-admit cells are no longer a single source: they pool the
# 0729 set below with a level-selected subset of pacing_only_0803.  Reproduce
# them with scripts/rq2_audit_pool.py, which owns the selection rule; the B1/B2
# blocks here remain the per-source views.
PUBLISHED_AUDIT = {
    ('12MP normal', 'Multi-frame'): (831, 16, +1.2, 4, 31, -3.0),
    ('12MP normal', 'Single-frame'): (842, 5, +0.7, 0, 35, -3.5),
    ('24MP memory', 'Multi-frame'): (892, 50, +2.6, 1, 53, -3.2),
    ('24MP memory', 'Single-frame'): (905, 27, +1.6, 3, 51, -3.4),
}

# The 0727 pool reproduces the published 12MP cells exactly and lands within two
# captures of the published 24MP cells once runs carrying a Capture Timeout are
# dropped, so that filter is applied to every enforced set for comparability.
CURRENT_ENFORCED = {
    '12MP normal': sorted(__import__('glob').glob(
        os.path.join(ML, 'data', '0727', '48U_metrics_12MP_normal_0727_*.xlsx'))),
    '24MP memory': sorted(__import__('glob').glob(
        os.path.join(ML, 'data', '0727', '48U_metrics_24MP_memory_0727_*.xlsx'))),
}


def drop_timeout_runs(dec):
    bad = {(d['book'], d['run']) for d in dec if d['timed_out']}
    return [d for d in dec if (d['book'], d['run']) not in bad]


def enforced_cells(paths, drop_to=False):
    dec, st = load_decisions(paths)
    if drop_to:
        dec = drop_timeout_runs(dec)
    out = {}
    for group in ('Multi-frame', 'Single-frame'):
        g = [d for d in dec if d['group'] == group]
        adm = [d for d in g if d['eff_admit']]
        miss = [d for d in adm if d['cost'] is None]
        wd = [d for d in adm if d['watchdog']]
        over = [d for d in adm if not d['watchdog'] and d['cost'] is not None
                and d['cost'] > d['budget']]
        unsafe = len(wd) + len(over)
        out[group] = (len(adm), len(g), len(adm) - unsafe - len(miss), unsafe,
                      len(wd), len(over))
    return out, st


def audit_cells(paths):
    dec, st = load_decisions(paths)
    D = deadline_of(paths)
    out = {}
    for group in ('Multi-frame', 'Single-frame'):
        g = [d for d in dec if d['group'] == group and d['cost'] is not None]
        feas = [d for d in g if not d['watchdog'] and d['cost'] <= d['budget']]
        unsf = [d for d in g if d['watchdog'] or d['cost'] > d['budget']]
        fa = sum(1 for d in feas if d['model_admit'])
        ua = sum(1 for d in unsf if d['model_admit'])
        marg = [100.0 * (d['budget'] - d['cost']) / D
                for d in feas if not d['model_admit']]
        over = [100.0 * (d['cost'] - d['budget']) / D
                for d in unsf if not d['model_admit']]
        out[group] = (fa, len(feas) - fa,
                      statistics.median(marg) if marg else None,
                      ua, len(unsf) - ua,
                      statistics.median(over) if over else None)
    return out, st, D


def compare():
    print('\n' + '#' * 104)
    print('# BLOCK 1  Controller-enforced runs')
    print('#   published : ML/data/0727/48U_metrics_<cond>_0727_*.xlsx, timeout-bearing runs dropped')
    print('#   new       : data/ablation_sampling/48U_metrics_<cond>_0803_{1,2}.xlsx  (Full arm)')
    print('#' * 104)
    new_e = ENFORCED['A3  data/ablation_sampling  Full (_1 + _2)']
    orig_e = ENFORCED['A2  data/ablation_original  Full (_1 + _2)']
    hdr = (f'{"condition":13s} {"group":13s} {"source":26s} {"admit rate":>18s} '
           f'{"successful":>18s} {"unsafe":>16s}')
    print(hdr)
    print('-' * len(hdr))
    for cond in ('12MP normal', '24MP memory'):
        rows = [('published table', None),
                ('recomputed 0727 pool', enforced_cells(CURRENT_ENFORCED[cond], True)[0]),
                ('new ablation_original', enforced_cells(orig_e[cond], True)[0]),
                ('new ablation_sampling', enforced_cells(new_e[cond], True)[0])]
        for group in ('Multi-frame', 'Single-frame'):
            for name, cells in rows:
                if cells is None:
                    a, n, s, u, tag = PUBLISHED_ENFORCED[(cond, group)]
                    w = o = None
                else:
                    a, n, s, u, w, o = cells[group]
                extra = '' if w is None else f' [wd {w}, C>B {o}]'
                print(f'{cond:13s} {group:13s} {name:26s} '
                      f'{pct(a, n):6.1f}% [{a:5d}/{n:5d}] '
                      f'{pct(s, a):6.1f}% [{s:5d}]   '
                      f'{pct(u, a):5.1f}% [{u}]{extra}')
            print()

    print('\n' + '#' * 104)
    print('# BLOCK 2  Always-admit model audit')
    print('#   published : the two sources below pooled, the 0803 one level-selected;')
    print('#               run scripts/rq2_audit_pool.py to reproduce that pooling')
    print('#   0729      : ML/data/0729_RQ2/48U_metrics_<cond>_0729_PacingOnly_{1,2}.xlsx')
    print('#   new       : data/ablation_sampling/48U_metrics_<cond>_pacing_only_0803.xlsx')
    print('#' * 104)
    hdr = (f'{"condition":13s} {"group":13s} {"source":26s} {"n":>5s} | '
           f'{"feas-admit":>13s} {"feas-skip":>13s} {"margin":>8s} | '
           f'{"unsafe-admit":>13s} {"unsafe-skip":>13s} {"overrun":>8s}')
    print(hdr)
    print('-' * len(hdr))
    for cond in ('12MP normal', '24MP memory'):
        cur = audit_cells(AUDIT['B1  current  0729 PacingOnly (_1 + _2)'][cond])[0]
        new = audit_cells(AUDIT['B2  data/ablation_sampling  pacing_only_0803'][cond])[0]
        for group in ('Multi-frame', 'Single-frame'):
            fa, fs, m, ua, us, o = PUBLISHED_AUDIT[(cond, group)]
            rows = [('published table', (fa, fs, m, ua, us, -o if o else o)),
                    ('recomputed 0729 audit', cur[group]),
                    ('new pacing_only_0803', new[group])]
            for name, c in rows:
                fa, fs, m, ua, us, o = c
                tot_f, tot_u = fa + fs, ua + us
                print(f'{cond:13s} {group:13s} {name:26s} {tot_f + tot_u:5d} | '
                      f'{fa:5d} ({pct(fa, tot_f):5.1f}%) {fs:5d} ({pct(fs, tot_f):5.1f}%) '
                      f'{("--" if m is None else f"+{m:.1f}%"):>8s} | '
                      f'{ua:5d} ({pct(ua, tot_u):5.1f}%) {us:5d} ({pct(us, tot_u):5.1f}%) '
                      f'{("--" if o is None else f"-{abs(o):.1f}%"):>8s}')
            print()


if __name__ == '__main__':
    only = sys.argv[1] if len(sys.argv) > 1 else None
    if only == 'compare':
        compare()
        sys.exit(0)
    for label, sets in ENFORCED.items():
        if only and not label.startswith(only):
            continue
        enforced_block(label, sets)
    for label, sets in AUDIT.items():
        if only and not label.startswith(only):
            continue
        audit_block(label, sets)
