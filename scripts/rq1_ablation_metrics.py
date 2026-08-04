"""RQ1(b) expanded ablation: 2 conditions x starting Lv3/Lv4 x 4 arms.

Follows docs/rq1-rq3-metrics-guide.md sections 3.2 (run reconstruction),
3.4 (inclusive percentiles), 3.5 (valid-run policy / dedup) and 4.3.
"""
import openpyxl, warnings, os, statistics
warnings.filterwarnings('ignore')

PAPER = r'C:/Users/sal_eunki/Desktop/SEIP-27'
ML    = r'C:/Users/sal_eunki/Desktop/ML'

ARMS = {
    ('12MP normal', 'No control'):     [],  # no accessible controller-off workbook
    ('12MP normal', 'Admission only'): [PAPER+'/data/48U_metrics_12MP_normal_0729_AdmitOnly_1.xlsx',
                                        PAPER+'/data/48U_metrics_12MP_normal_0729_AdmitOnly_2.xlsx'],
    ('12MP normal', 'Pacing only'):    [PAPER+'/data/48U_metrics_12MP_normal_0729_PacingOnly_1.xlsx',
                                        PAPER+'/data/48U_metrics_12MP_normal_0729_PacingOnly_2.xlsx'],
    ('12MP normal', 'Full'):           [ML+'/data/0803_FULL/SM-S948U_metrics_12MP_normal_0803.xlsx'],
    ('24MP memory', 'No control'):     [ML+'/data/0803_FULL/SM-S948U_metrics_24MP_memory_baseline_0803.xlsx'],
    ('24MP memory', 'Admission only'): [PAPER+'/data/48U_metrics_24MP_memory_0729_AdmitOnly_1.xlsx',
                                        PAPER+'/data/48U_metrics_24MP_memory_0729_AdmitOnly_2.xlsx'],
    ('24MP memory', 'Pacing only'):    [PAPER+'/data/48U_metrics_24MP_memory_0729_PacingOnly_1.xlsx',
                                        PAPER+'/data/48U_metrics_24MP_memory_0729_PacingOnly_2.xlsx'],
    ('24MP memory', 'Full'):           [ML+'/data/0803_FULL/SM-S948U_metrics_24MP_memory_0803.xlsx'],
}


# Predeclared per-cell exclusions, keyed (condition, arm, starting level) with
# the first-timeout indices of the sessions to drop.  Only one session is
# excluded: the 24MP Full Lv3 run that timed out at capture 11, which the table
# reports as 7/7.  Keeping the manifest here rather than in the reader means the
# retained form is one edit away, and that the criterion is visible.
#   Note the asymmetry this creates and that the manuscript must address: the
# criterion cannot be applied to all four arms, because No control and Pacing
# only would then have no runs left at 24MP.  It is a per-cell
# operator-invalidity call and has to be stated as one.
EXCLUDE = {('24MP memory', 'Full', 3): [11]}


def read_sheet(path, name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close(); return None, []
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    idx = {}
    for i, h in enumerate(hdr):
        if h and h not in idx:
            idx[h] = i
    rows = [r for r in it if any(c is not None for c in r)]
    wb.close()
    return idx, rows


def cell(r, i):
    return r[i] if (i is not None and len(r) > i) else None


def truthy(v):
    if v is None: return False
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return v != 0
    return str(v).strip().lower() in ('true', 'yes', '1', 'y')


def split_runs(rows, ppi):
    runs, cur, prev = [], [], None
    for r in rows:
        pp = cell(r, ppi)
        if cur and prev is not None and pp is not None and pp <= prev:
            runs.append(cur); cur = []
        cur.append(r)
        prev = pp
    if cur:
        runs.append(cur)
    return runs


def pctl_inc(vals, q):
    """Excel PERCENTILE.INC."""
    if not vals: return None
    s = sorted(vals)
    if len(s) == 1: return s[0]
    pos = (len(s) - 1) * q
    lo = int(pos); hi = min(lo + 1, len(s) - 1)
    return s[lo] + (s[hi] - s[lo]) * (pos - lo)


def load_arm(paths):
    """Return list of run dicts, deduplicated across workbooks."""
    seen, out = set(), []
    for p in paths:
        if not os.path.exists(p):
            print(f'  !! missing {p}'); continue
        ci, crows = read_sheet(p, 'Capture')
        runs = split_runs(crows, ci['ppSequenceId'])
        # per-run pacing totals from RQ3Summary, joined positionally (the
        # exporter applies the same run rule) and validated by shot count.
        si, srows = read_sheet(p, 'RQ3Summary')
        summ = srows if si else []
        for k, run in enumerate(runs):
            sig = tuple((cell(r, ci['captureIndex']), cell(r, ci['ppSequenceId']),
                         cell(r, ci['shotToShotTimeMs']),
                         cell(r, ci['draftSequenceDurationMs'])) for r in run)
            if sig in seen:
                continue
            seen.add(sig)
            pref = run[:30]
            lvl = cell(run[0], ci['firstNodeOverheatLevel'])
            to_pos = None
            for j, r in enumerate(pref):
                if truthy(cell(r, ci['isTimeout'])):
                    to_pos = j + 1; break
            def rate(col):
                vals = [1.0 if truthy(cell(r, ci[col])) else 0.0 for r in pref]
                return 100.0 * sum(vals) / len(vals) if vals else None
            ms = [1.0 if (truthy(cell(r, ci['bokehCompleted'])) and
                          truthy(cell(r, ci['filterCompleted']))) else 0.0 for r in pref]
            slack = []
            for r in pref:
                m = cell(r, ci['timeoutMarginMs'])
                if isinstance(m, (int, float)):
                    slack.append(float(m))
            # Lower-tail margin of THIS run.  Taking the percentile per run and
            # aggregating afterwards keeps long runs from dominating the cell,
            # which pooling every capture into one list did.
            slack_p5 = pctl_inc(slack, 0.05) if slack else None
            # Failure severity of THIS run: how far past the deadline the first
            # timeout went.  None when the run did not time out, so the cell
            # statistic is defined only over runs that actually failed.
            overrun = None
            if to_pos is not None:
                m = cell(pref[to_pos - 1], ci['timeoutMarginMs'])
                if isinstance(m, (int, float)):
                    overrun = -float(m)
            tot_delay = None; paced_pct = None
            if summ and k < len(summ):
                srow = summ[k]
                if cell(srow, si['shotCount']) in (len(run), len(run) - 1, len(run) + 1):
                    td = cell(srow, si['totalDelayMs'])
                    tot_delay = float(td) if isinstance(td, (int, float)) else None
                    pp = cell(srow, si['pacedPercent'])
                    paced_pct = float(pp) if isinstance(pp, (int, float)) else None
            out.append(dict(lvl=None if lvl is None else int(lvl), n=len(run),
                            reached30=len(run) >= 30, timeout_at=to_pos,
                            ms=100.0 * sum(ms) / len(ms) if ms else None,
                            m=rate('bokehCompleted'), s=rate('filterCompleted'),
                            slack=slack, slack_p5=slack_p5, overrun=overrun,
                            total_delay=tot_delay, paced=paced_pct,
                            src=os.path.basename(p)))
    return out


CTO = {}  # captureTimeoutMs per workbook, for slack normalization
for cond, arm in ARMS:
    for p in ARMS[(cond, arm)]:
        if not os.path.exists(p) or p in CTO: continue
        pi, prows = read_sheet(p, 'PacingReplay')
        if pi and 'captureTimeoutMs' in pi:
            vs = [float(cell(r, pi['captureTimeoutMs'])) for r in prows
                  if isinstance(cell(r, pi['captureTimeoutMs']), (int, float))]
            CTO[p] = statistics.median(vs) if vs else None

print('captureTimeoutMs per workbook:')
for k, v in CTO.items():
    print(f'  {os.path.basename(k):55s} {v}')
print()

DEADLINE = statistics.median([v for v in CTO.values() if v]) if CTO else None

hdr = (f"{'condition':13s} {'arm':15s} {'Lv':>2s} {'N':>3s} {'S30':>6s} "
       f"{'E':>3s} {'Med':>5s} {'M+S%':>6s} {'M%':>6s} {'slackP5%':>8s} "
       f"{'(p50)':>6s} {'(min)':>6s} {'nSurv':>5s} {'ovrP50%':>7s} {'ovrMs':>6s} "
       f"{'nTO':>3s} "
       f"{'SdelayP50s':>10s} {'paced%':>7s}")
print(hdr); print('-' * len(hdr))
print('slackP5% is the macro-average over surviving runs of each run\'s own P5;')
print('(p50)/(min) are the median and worst of the same per-run values, printed')
print('for the prose.  ovrP50% is the median overrun at the first timeout over')
print('the runs that failed.  Each is censored when its own run set is empty.')
print()

for (cond, arm), paths in ARMS.items():
    runs = load_arm(paths)
    for lv in (3, 4):
        g = [r for r in runs if r['lvl'] == lv]
        if not g:
            print(f'{cond:13s} {arm:15s} {lv:2d}   --  (no accessible source)')
            continue
        drop = EXCLUDE.get((cond, arm, lv))
        if drop:
            before = len(g)
            g = [r for r in g if r['timeout_at'] not in drop]
            print(f'  ** {cond} / {arm} / Lv{lv}: excluded {before - len(g)} '
                  f'session(s) timing out at {drop} (predeclared manifest)')
        n = len(g)
        s30 = sum(1 for r in g if r['reached30'] and r['timeout_at'] is None)
        tos = [r['timeout_at'] for r in g if r['timeout_at'] is not None]
        E = min(tos) if tos else None
        Med = statistics.median(tos) if len(tos) * 2 >= n else None  # censored otherwise
        ms = [r['ms'] for r in g if r['ms'] is not None]
        mm = [r['m'] for r in g if r['m'] is not None]
        # Slack P5 is reported only over the runs that completed 30 captures
        # with no Capture Timeout.  A truncated run contributes exactly one
        # negative margin -- the overrun that ended it -- so pooling every
        # capture of every run made the printed percentile a statistic of
        # failure severity whenever the mean surviving length fell below about
        # 20 captures, and made its sign depend on run length rather than on
        # deadline safety.  Failure severity is now its own column.
        surv = [r for r in g if r['reached30'] and r['timeout_at'] is None]
        sp5 = [r['slack_p5'] for r in surv if r['slack_p5'] is not None]
        nrm = lambda v: 100.0 * v / DEADLINE if (v is not None and DEADLINE) else None
        p5pct = nrm(sum(sp5) / len(sp5)) if sp5 else None      # macro-average
        p5p50 = nrm(statistics.median(sp5)) if sp5 else None
        p5min = nrm(min(sp5)) if sp5 else None
        ov = [r['overrun'] for r in g if r['overrun'] is not None]
        ovms = statistics.median(ov) if ov else None
        ovpct = nrm(ovms) if ov else None
        td = [r['total_delay'] for r in g if r['total_delay'] is not None]
        tdp50 = statistics.median(td) / 1000.0 if td else None
        pc = [r['paced'] for r in g if r['paced'] is not None]
        f = lambda v, d=1: ('--' if v is None else f'{v:.{d}f}')
        print(f'{cond:13s} {arm:15s} {lv:2d} {n:3d} {s30:2d}/{n:<3d} '
              f'{("--" if E is None else str(E)):>3s} {f(Med):>5s} '
              f'{f(sum(ms)/len(ms)) if ms else "--":>6s} '
              f'{f(sum(mm)/len(mm)) if mm else "--":>6s} '
              f'{f(p5pct):>8s} {f(p5p50):>6s} {f(p5min):>6s} {len(sp5):5d} '
              f'{f(ovpct):>7s} {f(ovms,0):>6s} {len(ov):3d} '
              f'{f(tdp50,2):>10s} {f(sum(pc)/len(pc)) if pc else "--":>7s}')
    print()
